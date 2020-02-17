#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  9 20:57:29 2020

@author: Mishaun_Bhakta
"""
import os, re, shutil

#sale parameters
state = "New Mexico"
stinitials = "NM"
date = "Feb 6, 2020"
bidder = '3'

#Navigate to energynet/govt sale and get sale page
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup

url = 'https://www.energynet.com/govt_listing.pl?sg=5196'

filepath = os.path.dirname(__file__)

#driver will be used based on operating system - windows or mac
try:
    driver = webdriver.Chrome(filepath + "/chromedriver.exe")
except:
    driver = webdriver.Chrome(filepath + "/chromedriver")

driver.implicitly_wait(30)
driver.get(url)

#storing html content in variable after reaching target sale page
salehtml = BeautifulSoup(driver.page_source, "html.parser")

def webscrape_presale(parsepage):
    '''This function will take a page and scrape its data for sale lot information
    '''

    #webscrape sale page
    serialnums = parsepage.find_all("span", "lot-name")
    serialnums = [i.text for i in serialnums]
    
    #storing all data from tag 'td's with clas name "lot-legal
    legalinfo = parsepage.find_all("td", "lot-legal")
    
    #initializing empty arrays
    acres = []
    desc = []
    county = []
    
    for item in legalinfo:
        county.append(item.contents[0].text)
        desc.append(item.contents[1].text)
        #getting acres by splitting at : and blankspace to get string of numerical value - taking out a comma if above 1000 in order to convert to float
        acres.append(float(re.split(":\W",item.contents[2])[1].replace(',','')))

    return acres, desc, county, serialnums

acres, descriptions, counties, serials = webscrape_presale(salehtml)

#Open sale template and update insert information from webscrape
import openpyxl

def fillexcel():
    '''
    This function will take scraped (global) values for lots and insert into sale spreadsheet and also download sale shapefile and move it to directory of this script file
    '''    
    
    #opening template sale notebook for modifications
    #preserving vba to keep formatting of workbook preserved - also keeping formulas 
    wb = openpyxl.load_workbook("BLM Sale Notes Template.xlsm", keep_vba = True)
    sheet = wb.active
    
    sheet["B6"] = "BLM {} {} Sale Notes".format(stinitials, date)
    #inserting values from webscrape into spreadsheet
    for i in range(0,len(serials)):
        sheet.cell(row = 8+i, column = 2, value = serials[i])
        sheet.cell(row = 8+i, column = 6, value = acres[i])
        sheet.cell(row = 8+i, column = 7, value = counties[i])
        sheet.cell(row = 8+i, column = 8, value = descriptions[i])
    
    wb.save("BLM {} {} Sale Notes.xlsm".format(stinitials, date))
    wb.close()
    
    #clicking link of where shapefile is stored on sale page    
    driver.find_element_by_link_text("GIS Data WGS84").click()
    driver.find_element_by_link_text("Notice of Competitive Oil and Gas Internet-Based Lease Sale").click()
    
    #getting list of filenames in downloads
    try:
        downloaddir = "/Users/Mishaun_Bhakta/Downloads/"
        downloads = os.listdir(downloaddir)
    except:
        downloaddir = "C:/Users/mishaun/Downloads/"
        downloads = os.listdir(downloaddir) 
        
    #pattern will find downloaded file name of shapefile
    pattern = "BLM"+ stinitials + "\S*.zip"
    
    #searching through filenames in downlaods folder
    finds = []
    for file in downloads:
        if re.findall(pattern, file):
            finds.append(file)
            break
        
    #moving file from downloads folder to directory of this script file - then renaming it to a cleaner name
    shutil.copy(downloaddir + finds[0], filepath)
    os.rename(finds[0], "BLM " + stinitials + " " + date + " Shapefile." + finds[0].split(".")[1])

#checking to see whether or not excel file already exists - if it does it'll prevent overwriting of changes
if os.path.exists(filepath+ "/" + "BLM {} {} Sale Notes.xlsm".format(stinitials, date)):
    print("File already exists - Preventing overwrite of changes in excel file")
else:
    fillexcel()
    
bidtags = salehtml.find_all("td", "lot-bid")
ourwinnings = {}

for i in range (0,len(bidtags)):
    textCont = bidtags[i].text
    
    #extracting bidder number from lot bid tags - try statement prevents break if no bids were received
    try:
        winBidder = re.findall('#\d+', textCont)[0].replace("#",'')
    except:
        print("no bids received for parcel: " + serials[i])
        winBidder = type(None)
        pass
    
    #if we won the bid, then capture the winning bid amount
    if bidder == winBidder:
        winAmount = re.findall('\$\d+', textCont)[0]
        winAmount = winAmount.replace('$','')
    
        ourwinnings[i] = winAmount

def fillwinnings():
    '''This function will take ourwinnings dictionary and add values to created spreadsheet
    '''
    
    #### insert our winnings into sale spreadsheet
    wb = openpyxl.load_workbook("BLM {} {} Sale Notes.xlsm".format(stinitials, date), keep_vba = True)
    sheet = wb.active
    
    for i in range(0,len(ourwinnings)):
        #row 8 is the starting row for parcels in teh spreadsheet, inserting data relative to 8th row by adding parcel number of sale
        sheet.cell(row = 8 + list(ourwinnings.keys())[i], column = 17, value = ourwinnings[list(ourwinnings.keys())[i]])
        sheet.cell(row = 8 + list(ourwinnings.keys())[i], column = 16, value = 'Y')
    
    wb.save("BLM {} {} Sale Notes.xlsm".format(stinitials, date))
    wb.close()

#create dataframe for completed sale sheet
import pandas as pd


#use pdf reader to fill in form
# conda install -c conda-forge pdfrw
import pdfrw

#copied code and function from article: https://bostata.com/how-to-populate-fillable-pdfs-with-python/
##############################################################################
ANNOT_KEY = '/Annots'
ANNOT_FIELD_KEY = '/T'
ANNOT_VAL_KEY = '/V'
ANNOT_RECT_KEY = '/Rect'
SUBTYPE_KEY = '/Subtype'
WIDGET_SUBTYPE_KEY = '/Widget'

def write_fillable_pdf(input_pdf_path, output_pdf_path, data_dict):
    template_pdf = pdfrw.PdfReader(input_pdf_path)
    annotations = template_pdf.pages[0][ANNOT_KEY]
    for annotation in annotations:
        if annotation[SUBTYPE_KEY] == WIDGET_SUBTYPE_KEY:
            if annotation[ANNOT_FIELD_KEY]:
                key = annotation[ANNOT_FIELD_KEY][1:-1]
                if key in data_dict.keys():
                    annotation.update(
                        pdfrw.PdfDict(V='{}'.format(data_dict[key]))
                    )
    pdfrw.PdfWriter().write(output_pdf_path, template_pdf)
##############################################################


def wonlotsDF():
    #using openpyxl in order to read formulated values from spreadsheet
    # NOTE: have to manually open excel and save sheet for formulated cells to read after filling in values
    data_onlyWB = openpyxl.load_workbook("BLM {} {} Sale Notes.xlsm".format(stinitials, date), data_only = True, keep_vba = True)
    dataSheet = data_onlyWB.active
    
    df = pd.DataFrame(dataSheet.values)
    
    
    #slicing the dataframe to get only relevant data
    df = df.iloc[6:,1:25]
    #setting columns to first row of dataframe
    df.columns = df.iloc[0]
    #dropping the repeated row with column names
    df = df.drop(index =[6])
    
    #filtering data frame with values only won by magnum
    wonlotsdf = df[df["Magnum Won (Y/N)"] == 'Y']
    return wonlotsdf


def createBidSheets():
    '''
    This function will take a template pdf and generate pdf's based on wonlots dataframe
    '''
    #calling funciton wonlotsDF in ordre for bid sheets to be created 
    wonlotsdf = wonlotsDF()

    templatePDF = 'bidsheet template.pdf'
    
    for i in range(0,len(wonlotsdf.index)):
    
        OutputPath = filepath +"/Bid Sheets/" + wonlotsdf.iloc[i]["Serial numbers"] + " Bid Sheet.pdf"
        
        fields = {
                "State": stinitials,
                "Date of Sale": date,
                'Check Box for Oil and Gas' : "x",
                "Oil and Gas/Parcel No" : wonlotsdf.iloc[i]["Serial numbers"],
                "TOTAL BID FOR Oil and Gas Lease" : wonlotsdf.iloc[i]["Total Bid (Number on BLM Bid Sheet)"],
                "PAYMENT SUBMITTED WITH BID for Oil and Gas" : wonlotsdf.iloc[i]["Min Due"],
                "Print or Type Name of Lessee" : "R&R Royalty, LTD",
                "Address of Lessee": "500 N Shoreline Blvd, Ste 322",
                "City" : "Corpus Christi",
                "State_2": "TX",
                "Zip Code" : "78401"
                }
        
        write_fillable_pdf(templatePDF, OutputPath, fields)